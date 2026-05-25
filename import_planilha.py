#!/usr/bin/env python3
"""
import_planilha.py — Importa jogadores da planilha Excel para o Fut Segunda

Uso:
    python import_planilha.py planilha.xlsx           # importa de verdade
    python import_planilha.py planilha.xlsx --dry-run  # só confere, não sobe

Abas esperadas:
    MENSALISTAS — col A: nome do jogador, col B: valor pago
    AVULSOS     — col A: nome do jogador, col B: valor pago

    Nomes com "ISENTO" são marcados como isentos no sistema e mantêm o nome igual.
"""

import sys, json, re, getpass, unicodedata
import openpyxl

# ═══════════════════════════════════════════════════════
#  CONFIGURAÇÃO
# ═══════════════════════════════════════════════════════
API_URL    = "https://fut.barleiseca.com.br/api/api.php"
API_KEY    = "fut2-minha-chave-secreta"
FEE_MENSAL = 70
FEE_AVULSO = 20
DATA_REF   = "2025-01-01"   # data de referência para pagamentos históricos

# ═══════════════════════════════════════════════════════
#  ESTADO
# ═══════════════════════════════════════════════════════
_next_id = [1]
players  = {}   # norm_nome → dict jogador

def novo_id():
    v = _next_id[0]; _next_id[0] += 1; return v

# ═══════════════════════════════════════════════════════
#  UTILITÁRIOS
# ═══════════════════════════════════════════════════════
def norm(s):
    s = unicodedata.normalize('NFKD', str(s)).encode('ascii', 'ignore').decode()
    return re.sub(r'\s+', ' ', s).strip().lower()

def limpo(v):
    return str(v).strip() if v is not None else ''

def parse_valor(v):
    s = re.sub(r'[R$\s]', '', str(v) if v else '').replace('.', '').replace(',', '.')
    try:    return float(s)
    except: return 0.0

IGNORAR = {
    'total', 'despesas', 'receitas', 'saldo', 'pgto', 'mensalistas',
    'mensal', 'avulso', 'avulsos', 'jantas', 'luz', 'valor', 'nome',
}

def eh_nome(v):
    s = limpo(v)
    if not s or s == '-': return False
    if re.match(r'^R\$', s): return False
    if re.match(r'^[\d,.]+$', s): return False
    if norm(s) in IGNORAR: return False
    return True

# ═══════════════════════════════════════════════════════
#  JOGADORES
# ═══════════════════════════════════════════════════════
def get_ou_criar(nome_raw, is_regular, is_isento=False):
    nome = limpo(nome_raw)
    key  = norm(nome)
    # detecta isento pelo nome ou pelo flag explícito
    is_isento = is_isento or ('isento' in key)

    if key not in players:
        players[key] = {
            'id':         novo_id(),
            'name':       nome.title(),
            'position':   'Meio',
            'attributes': {'physical': 65, 'tactical': 60, 'technical': 70},
            'overall':    64,
            'isRegular':  is_regular,
            'isIsento':   is_isento,
            'monthlyFee': FEE_MENSAL if is_regular else FEE_AVULSO,
            'balance':    0.0,
            'payments':   [],
            'dinnerDebt': 0,
            'lastRating': None,
        }
    elif is_isento:
        # se já existe mas agora sabemos que é isento, atualiza o flag
        players[key]['isIsento'] = True
    return players[key]

def add_pagamento(jogador, tipo, valor):
    if valor <= 0: return
    jogador['payments'].append({'type': tipo, 'amount': valor, 'date': DATA_REF})
    jogador['balance'] += valor

# ═══════════════════════════════════════════════════════
#  PARSER DE ABA
# ═══════════════════════════════════════════════════════
def processar_aba(ws, is_regular):
    tipo = 'Mensalidade' if is_regular else 'Jogo Avulso'
    conta = 0
    for row in ws.iter_rows(values_only=True):
        nome = limpo(row[0] if row else None)
        if not eh_nome(nome):
            continue
        valor     = parse_valor(row[1] if len(row) > 1 else None)
        is_isento = (valor == 0)   # valor 0 na planilha = jogador isento
        jog       = get_ou_criar(nome, is_regular, is_isento)
        add_pagamento(jog, tipo, valor)
        conta += 1
    return conta

# ═══════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════
def upload_snapshot(snapshot):
    """Autentica e envia o snapshot ao servidor."""
    try:
        import requests as req
    except ImportError:
        print("ERRO: Instale: pip install openpyxl requests")
        sys.exit(1)

    # Headers que imitam o navegador para não ser bloqueado pelo Mod_Security do Hostgator
    BASE_HEADERS = {
        'X-Api-Key':    API_KEY,
        'Content-Type': 'application/json',
        'Accept':       'application/json',
        'User-Agent':   'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/124.0 Safari/537.36',
    }

    print("  Credenciais do sistema:")
    usuario = input("  Usuario [admin]: ").strip() or 'admin'
    senha   = getpass.getpass("  Senha: ")

    resp = req.post(
        API_URL + "?action=login",
        json={'username': usuario, 'password': senha},
        headers=BASE_HEADERS,
    )
    if not resp.ok:
        print(f"ERRO: Login falhou ({resp.status_code}): {resp.text[:300]}")
        sys.exit(1)
    token = resp.json().get('token')
    print("  Login OK\n")

    auth_headers = {**BASE_HEADERS, 'X-Auth-Token': token}

    existing = req.get(API_URL, headers=auth_headers).json()
    qtd = len(existing.get('players', []))
    if qtd > 0:
        print(f"  AVISO: O sistema ja tem {qtd} jogador(es).")
        ok = input("  Sobrescrever tudo? [s/N]: ").strip().lower()
        if ok != 's':
            print("  Cancelado."); sys.exit(0)
        print()

    r = req.post(API_URL, json=snapshot, headers=auth_headers)
    if r.ok:
        print("  Jogadores importados com sucesso!")
        print("  Acesse o sistema e confira o painel de jogadores.\n")
    else:
        print(f"  ERRO ao importar: {r.status_code} - {r.text[:300]}")


def main():
    dry_run     = '--dry-run'     in sys.argv
    upload_seed = '--upload-seed' in sys.argv
    args        = [a for a in sys.argv[1:] if not a.startswith('--')]

    # Modo: sobe seed.json existente sem ler planilha
    if upload_seed:
        seed_file = args[0] if args else 'seed.json'
        print(f"\nCarregando {seed_file}...")
        with open(seed_file, encoding='utf-8') as f:
            snapshot = json.load(f)
        qtd = len(snapshot.get('players', []))
        print(f"  {qtd} jogadores encontrados no seed.")
        upload_snapshot(snapshot)
        return

    if not args:
        print("Uso:")
        print("  python import_planilha.py planilha.xlsx [--dry-run]")
        print("  python import_planilha.py --upload-seed [seed.json]")
        sys.exit(1)

    caminho = args[0]
    print(f"\nLendo: {caminho}\n")

    wb = openpyxl.load_workbook(caminho, data_only=True)

    # Localizar abas pelo nome
    aba_mensal = next((s for s in wb.sheetnames if 'mensal' in norm(s)), None)
    aba_avulso = next((s for s in wb.sheetnames if 'avulso' in norm(s)), None)

    if not aba_mensal and not aba_avulso:
        print("❌ Nenhuma aba 'MENSALISTAS' ou 'AVULSOS' encontrada.")
        print(f"   Abas disponíveis: {', '.join(wb.sheetnames)}")
        sys.exit(1)

    qtd_m = qtd_a = 0
    if aba_mensal:
        qtd_m = processar_aba(wb[aba_mensal], is_regular=True)
        print(f"  OK {aba_mensal}: {qtd_m} mensalista(s) lido(s)")
    if aba_avulso:
        qtd_a = processar_aba(wb[aba_avulso], is_regular=False)
        print(f"  OK {aba_avulso}: {qtd_a} avulso(s) lido(s)")

    if not players:
        print("❌ Nenhum jogador encontrado.")
        sys.exit(1)

    # ─ Montar snapshot ─────────────────────────────────
    snapshot = {
        '_id':           _next_id[0] + 1,
        'players':       list(players.values()),
        'fees':          {'mensal': FEE_MENSAL, 'avulso': FEE_AVULSO},
        'attendances':   [],
        'results':       [],
        'teamHistory':   [],
        'dinnerHistory': [],
        'expenses':      [],
        'config': {
            'logo':           None,
            'teamName':       'Fut Segunda',
            'tabTitle':       'Fut Segunda — Manager',
            'initialBalance': 0,
        },
    }

    # ─ Resumo ──────────────────────────────────────────
    mensalistas = [p for p in players.values() if p['isRegular']]
    avulsos     = [p for p in players.values() if not p['isRegular']]
    isentos     = [p for p in players.values() if p['isIsento']]
    total_rec   = sum(p['balance'] for p in players.values())

    print()
    print("=" * 48)
    print("  RESUMO DA IMPORTACAO")
    print("=" * 48)
    print(f"  Mensalistas : {len(mensalistas)}")
    print(f"  Avulsos     : {len(avulsos)}")
    print(f"  Isentos     : {len(isentos)}")
    total_str = f"{total_rec:.2f}".replace('.', ',')
    print(f"  Total pago  : R$ {total_str}")
    print()
    if mensalistas:
        print("  Mensalistas:", ', '.join(p['name'] for p in mensalistas))
    if avulsos:
        print("  Avulsos    :", ', '.join(p['name'] for p in avulsos))
    if isentos:
        print("  Isentos    :", ', '.join(p['name'] for p in isentos))
    print()

    if dry_run:
        with open('seed.json', 'w', encoding='utf-8') as f:
            json.dump(snapshot, f, ensure_ascii=False, indent=2)
        print("  [dry-run] seed.json gerado - nada enviado ao servidor.\n")
        return

    upload_snapshot(snapshot)

if __name__ == '__main__':
    main()
